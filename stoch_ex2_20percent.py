# Working 2
import numpy as np
from pyomo.environ import *
#from pyomo import *
#import pyomo.environ as pyo
from pyomo.opt import SolverFactory
from pyomo.contrib.iis import *
import openpyxl
import random
from bayes_opt import BayesianOptimization
from bayes_opt import UtilityFunction
#from bayes_opt import acquisition

import gym
import copy
from gym import spaces
import numpy as np
class Blending_env(gym.Env):
    def __init__(self,config=None):
        n_streams = 1
        n_output = 1
        n_properties = 1
        n_blenders = 1
        n_periods = 6
        periods = range(0,n_periods)
        points = range(0,n_periods+1)
        inv_capacity = {'streams':np.ones(n_streams)*1000,'blender':np.ones(n_blenders)*1000,'product':np.ones(n_output)*1000}
        supply = {}
        for j in periods:
            for i in range(0,n_streams):
                supply[i,j] = 0
        property = {}
        for j in range(0,n_properties):
            for i in range(0,n_streams):
                property[i,j] = 0
        lb_prop = {}
        for j in range(0,n_properties):
            for i in range(0,n_output):
                lb_prop[i,j] = 0
        ub_prop = {}
        for j in range(0,n_properties):
            for i in range(0,n_output):
                ub_prop[i,j] = 0
        Fixed_cost = {}
        for i in range(0,n_streams):
            for j in range(0,n_blenders):
                Fixed_cost[i,j,'sj'] = 0
        for i in range(0,n_blenders):
            for j in range(0,n_output):
                Fixed_cost[i,j,'jp'] = 0
        for i in range(0,n_blenders):
            for j in range(0,n_blenders):
                Fixed_cost[i,j,'jj'] = 0
        
        Variable_cost = {}
        for i in range(0,n_streams):
            for j in range(0,n_blenders):
                Variable_cost[i,j,'sj'] = 0
        for i in range(0,n_blenders):
            for j in range(0,n_output):
                Variable_cost[i,j,'jp'] = 0
        for i in range(0,n_blenders):
            for j in range(0,n_blenders):
                Variable_cost[i,j,'jj'] = 0
        demand_price = {}
        for i in range(0,n_output):
            demand_price[i] = 0
        
        self.default_config = {'n_streams':n_streams,'n_output':n_output,'n_properties':n_properties,'n_blenders':n_blenders,'n_periods':n_periods,
                               'supply':supply,'demand_price':demand_price,'property':property,'lb_prop':lb_prop,'ub_prop':ub_prop,'inv_capacity':inv_capacity,
                               'Fixed_cost' : Fixed_cost,'Variable_cost':Variable_cost}
        
        if config:
            self.default_config.update(config)
        #print(self.default_config)
        
        # Dynamically set attributes based on the configuration
        #for key, value in self.default_config.items():
            #setattr(self, key, value)
        n_streams = self.default_config['n_streams']
        n_blenders = self.default_config['n_blenders']
        n_output = self.default_config['n_output']
        n_properties = self.default_config['n_properties']
        self.current_step = 0
        supply_list = []
        for i in range(0,self.default_config['n_streams']):
            supply_list.append(self.default_config['supply'][i,1])
        self.observation_space_dict = spaces.Dict({
            'I_s': spaces.Box(low=0, high=1000, shape=(n_streams,), dtype=np.float32),
            'I_j': spaces.Box(low=0, high=1000, shape=(n_blenders,), dtype=np.float32),
            'I_p': spaces.Box(low = 0, high=100, shape=(n_output,), dtype=np.float32),
            'C_j': spaces.Box(low=0, high=10, shape=(n_properties,n_blenders,), dtype=np.float32),
            'supply': spaces.Box(low=0, high=100, shape=(n_streams,), dtype=np.float32)

        })

        self.original_shapes = {}
        self.original_lows = {}
        self.original_highs = {}
        self.total_elements = 0

        for key in self.observation_space_dict.spaces:
            space = self.observation_space_dict.spaces[key]
            shape = space.shape
            low = space.low
            high = space.high
            size = np.prod(shape)
            self.original_shapes[key] = shape
            self.original_lows[key] = low
            self.original_highs[key] = high
            self.total_elements += size

        self.observation_space = spaces.Box(
            low=np.concatenate([self.original_lows[key].flatten() for key in self.observation_space_dict.spaces]),
            high=np.concatenate([self.original_highs[key].flatten() for key in self.observation_space_dict.spaces]),
            dtype=np.float32
        )
        self.state = np.concatenate([np.zeros(n_streams), np.zeros(n_blenders), np.zeros(n_output),np.zeros(n_properties*n_blenders),np.array(supply_list)])
        
        #print(self.observation_space)
        self.action_space = self.create_action_space()
    def create_action_space(self):
        max_fsj = np.ones((self.default_config['n_streams'],self.default_config['n_blenders']))
        for s in range(0,self.default_config['n_streams']):
            for j in range(0,self.default_config['n_blenders']):
                max_fsj[s,j] = 50
        max_fjp = np.ones((self.default_config['n_blenders'],self.default_config['n_output']))
        for s in range(0,self.default_config['n_blenders']):
            for j in range(0,self.default_config['n_output']):
                #if(self.default_config['n_blenders']==1):
                max_fjp[s,j] = 50
        max_fjj = np.ones((self.default_config['n_blenders'],self.default_config['n_blenders']))
        #print(np.shape(max_fjj))
        #print(max_fjj)
        for s in range(0,self.default_config['n_blenders']):
            for j in range(0,self.default_config['n_blenders']):
                if(s==j):
                    max_fjj[s,j] = 0
                else:
                    max_fjj[s,j] = 50
        F_s_j_space = spaces.Box(low=0, high=max_fsj, shape=(self.default_config['n_streams'],self.default_config['n_blenders'],), dtype=np.float32)
        #F_j_j_space = spaces.Box(low=0, high=max_fjj, shape=(self.default_config['n_blenders'],self.default_config['n_blenders'],), dtype=np.float32)
        F_j_p_space = spaces.Box(low=0, high=max_fjp, shape=(self.default_config['n_blenders'],self.default_config['n_output'],), dtype=np.float32)
        pi_p_space = spaces.Box(low=0, high=100, shape=(self.default_config['n_output'],), dtype=np.float32)

        # Flatten the low and high bounds
        low = np.concatenate([F_s_j_space.low.flatten(),  F_j_p_space.low.flatten(),pi_p_space.low.flatten()])
        high = np.concatenate([F_s_j_space.high.flatten(), F_j_p_space.high.flatten(),pi_p_space.high.flatten()])

        # Create the flattened Box space
        flattened_action_space = spaces.Box(low=low, high=high, dtype=np.float32)
        
        
        return flattened_action_space

    def reset(self):
        supply_list = []
        for i in range(0,self.default_config['n_streams']):
            supply_list.append(self.default_config['supply'][i,1])
        self.state = np.concatenate([np.zeros(self.default_config['n_streams']), np.zeros(self.default_config['n_blenders']), np.zeros(self.default_config['n_output']),np.zeros(self.default_config['n_properties']*self.default_config['n_blenders']),np.array(supply_list)])
        self.current_step = 0
        return self.state
    def flatten_observation(self, obs):
        """Flatten the observation dictionary into a single array."""
        flattened = []
        for key in self.observation_space_dict.spaces:
            #print(obs)
            flattened.append(obs[key].flatten())
        return np.concatenate(flattened)

    def unflatten_observation(self, flattened_obs):
        """Unflatten the observation array back into the dictionary format."""
        offset = 0
        unflattened = {}
        for key in self.observation_space_dict.spaces:
            shape = self.original_shapes[key]
            size = np.prod(shape)
            unflattened[key] = flattened_obs[offset:offset + size].reshape(shape)
            offset += size
        return unflattened

    def step(self, actions):
        old_state = copy.deepcopy(self.state)
        old_state_dict = self.unflatten_observation(old_state)
        I_s_old = old_state_dict['I_s']
        I_j_old = old_state_dict['I_j']
        I_p_old = old_state_dict['I_p']
        C_j_old = old_state_dict['C_j']
        n_streams = self.default_config['n_streams']
        n_blenders = self.default_config['n_blenders']
        n_output = self.default_config['n_output']
        n_properties = self.default_config['n_properties']
       
        num_fsj = self.default_config['n_streams'] * self.default_config['n_blenders']
        num_fjp = self.default_config['n_blenders'] * self.default_config['n_output']
        #print(actions)
        #print(actions)
        F_s_j_action = actions[:num_fsj].reshape(self.default_config['n_streams'], self.default_config['n_blenders'])
        #F_j_j_action = actions[num_fsj:num_fsj + num_fjj].reshape(self.default_config['n_blenders'], self.default_config['n_blenders'])
        F_j_p_action = actions[num_fsj:num_fsj + num_fjp].reshape(self.default_config['n_blenders'], self.default_config['n_output'])
        pi_p_action = actions[num_fsj + num_fjp:].reshape(self.default_config['n_output'])
        action = {
        'F_s_j': F_s_j_action,
        'F_j_p': F_j_p_action,
        'pi_p':pi_p_action
        }
        penalty = 0
        new_state = copy.deepcopy(old_state_dict)
        for s in range(0,n_streams):
            new_state['I_s'][s] = I_s_old[s]+self.default_config['supply'][s,self.current_step+1]-sum(action['F_s_j'][s,j] for j in range(0,n_blenders))
            if(new_state['I_s'][s]<0):
                penalty = penalty+max(0,-new_state['I_s'][s])**2
        for j in range(0,n_blenders):
            new_state['I_j'][j] = I_j_old[j]+sum(action['F_s_j'][s,j] for s in range(0,n_streams))-sum(action['F_j_p'][j,p] for p in range(0,n_output))
            if(new_state['I_j'][j]<0):
                penalty = penalty+max(0,-new_state['I_j'][j])**2
        for p in range(0,n_output):
            new_state['I_p'][p] = I_p_old[p]-action['pi_p'][p]+sum(action['F_j_p'][j,p] for j in range(0,n_blenders))
            if(new_state['I_p'][p]<0):
                penalty = penalty+max(0,-new_state['I_p'][p])**2
        for j in range(0,n_blenders):
            for q in range(0,n_properties):
                if(new_state['I_j'][j]<=10**(-10)):
                    new_state['C_j'][q,j] = 0
                    #new_state['I_j'][j] = 0

                else:
                    new_state['C_j'][q,j] = (I_j_old[j]*C_j_old[q,j]+sum(action['F_s_j'][s,j]*self.default_config['property'][s,q] for s in range(0,n_streams))-sum(action['F_j_p'][j,p]*C_j_old[q,j] for p in range(0,n_output)))/new_state['I_j'][j]
                    
        self.current_step += 1
        supply_list = []
        for i in range(0,self.default_config['n_streams']):
            supply_list.append(self.default_config['supply'][i,self.current_step])

        new_state['supply'] = np.array(supply_list)
        done = self.current_step >= self.default_config['n_periods']
        tot_fixed_cost = 0
        tot_fixed_cost+= sum(self.default_config['Fixed_cost'][s,j,'sj']*(action['F_s_j'][s,j]!=0) for s in range(0,n_streams) for j in range(0,n_blenders))
        #tot_fixed_cost+= sum(self.default_config['Fixed_cost'][j,k,'jj']*(action['F_j_j'][j,k]!=0) for k in range(0,n_blenders) for j in range(0,n_blenders))
        tot_fixed_cost+= sum(self.default_config['Fixed_cost'][j,p,'jp']*(action['F_j_p'][j,p]!=0) for p in range(0,n_output) for j in range(0,n_blenders))
        tot_variable_cost = 0
        tot_variable_cost+= sum(self.default_config['Variable_cost'][s,j,'sj']*(action['F_s_j'][s,j]) for s in range(0,n_streams) for j in range(0,n_blenders))
        #tot_variable_cost+= sum(self.default_config['Variable_cost'][j,k,'jj']*(action['F_j_j'][j,k]) for k in range(0,n_blenders) for j in range(0,n_blenders))
        tot_variable_cost+= sum(self.default_config['Variable_cost'][j,p,'jp']*(action['F_j_p'][j,p]) for p in range(0,n_output) for j in range(0,n_blenders))
        
        for j in range(0,n_blenders):
            for p in range(n_output):
                if(I_j_old[j]!=0 and action['F_j_p'][j,p]!=0):
                    for q in range(0,n_properties):
                        if(C_j_old[q,j]<self.default_config['lb_prop'][p,q]):
                            #penalty = penalty+100*np.exp(100*(self.default_config['lb_prop'][p,q]-C_j_old[q,j]))
                            penalty = penalty+10**5
                        if(C_j_old[q,j]>self.default_config['ub_prop'][p,q]):
                            #penalty = penalty+100*np.exp(100*(C_j_old[q,j]-self.default_config['ub_prop'][p,q]))
                            penalty = penalty+10**5
                            #print("b")
        for j in range(0,n_blenders):
            for p in range(n_output):
                for s in range(0,n_streams):
                    if(action['F_j_p'][j,p]>0 and action['F_s_j'][s,j]>0):
                        print("yes")
                        penalty = penalty+10**5
                
        
        
        reward = -tot_fixed_cost-tot_variable_cost-penalty+sum(action['pi_p'][p]*self.default_config['demand_price'][p] for p in range(n_output))
        print(action['pi_p'])
        print(penalty)
        self.state = self.flatten_observation(new_state)
        return self.state, reward, done, {}
    def render(self, mode='human'):
        print(f"Step: {self.current_step}, State: {self.unflatten_observation(self.state)}")
        #print("yes")

    def close(self):
        pass



# Load an existing workbook or create a new one
file_path = "stoch_data1_test.xlsx"
try:
    # Try to load an existing workbook
    workbook = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    # If it doesn't exist, create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()

# Select the active sheet (or use workbook.create_sheet() to create a new one)
sheet = workbook.active

row_data = ["param1", "param2", "param3", "param4","param5","Obj"]

# Write the row data to a specific row (e.g., row 1)
sheet.append(row_data)



    
    
    

    


def instance_1(x1,x2,x3,x4):
    param_val = [0,0,0,0,0,0,0,0,0]
    param_val[0] = x1
    param_val[1] = x2
    param_val[2] = x3
    param_val[3] = 1
    param_val[4] = 1
    param_val[5] = 1
    param_val[6] = 1
    param_val[7] = 1
    param_val[8] = 1
    n_streams = 3
    n_output = 3
    n_properties = 3
    n_blenders = 4
    n_periods = 4
    mu, sigma = 0, 0.01 
    streams = range(n_streams)
    output = range(n_output)
    properties = range(n_properties)
    blenders = range(n_blenders)
    periods = range(1,n_periods+1)
    points = range(0,n_periods+1)
    periods_mb = range(1,n_periods+1)
    inv_capacity = {'streams':np.ones(n_streams)*100,'blender':np.ones(n_blenders)*100,'product':np.ones(n_output)*100}
    
    #var = 0.001
    
    supply = {}
    for j in periods:
        for i in range(0,n_streams):
            if(j==1):
                supply[i,j] = 200+np.random.normal(mu, 40)
            elif(j==2):
                supply[i,j] = 400+np.random.normal(mu, 80)
            elif(j==3):
                supply[i,j] = 600+np.random.normal(mu, 120)
            else:
                supply[i,j] = 0


    
    print(supply)
    demand_price = {}
    for i in range(0,n_output):
        demand_price[i] = 60+np.random.normal(mu, 12)
    property = {}
    #prop_lis = [[0,0,1,0,0.92,0.86,0],[0,0,0.96,0,0.91,0.71,0],[0,0,0.85,0,0.8,0.85,0],[0.33,0.07,0.81,0.52,0.84,0.97,1],[0.24,0.13,0.87,1,0.9,0.91,0.16],[0.04,0.08,0.7,0,0.66,0.83,0.03],[1,1,0.97,0,1,1,0]]
    for j in range(0,n_properties):
        for i in range(0,n_streams):
            property[i,j] = max(0,1.2*(i+1)*(j+1)/((n_properties+1)*(n_streams+1)))
            #property[i,j] = max(0,i*j/(n_properties*n_streams))
    #lb_prop_lis = [[0,0,0],[0,0,0],[0.87,0.91,0.93],[0,0,0],[0.89,0.92,0.95],[0.89,0.89,0.89],[0,0,0]]
    lb_prop = {}
    for j in range(0,n_properties):
        for i in range(0,n_output):
            #lb_prop[i,j] = max(0,(i*j)/(n_properties*n_output)+np.random.normal(mu, sigma/10))
            lb_prop[i,j] = max(0,(i+1)*(j+1)/((n_properties+1)*(n_output+1)))
    #ub_prop_lis = [[0.8,0.8,0.8],[0.79,0.79,0.79],[1,1,1],[0.9,0.9,0.9],[1,1,1],[1,1,1],[0.21,0.21,0.21]]
    ub_prop = {}
    for j in range(0,n_properties):
        for i in range(0,n_output):
            #ub_prop[i,j] = max(0,max(1,max(0,(2*i*j)/(n_properties*n_output)))+np.random.normal(mu, sigma/10))
            ub_prop[i,j] = max(0,max(1,max(0,(2*(i+1)*(j+1)/((n_properties+1)*(n_output+1))))))
    Fixed_cost = {}
    for i in range(0,n_streams):
        for j in range(0,n_blenders):
            Fixed_cost[i,j,'sj'] = 20+np.random.normal(mu, 4)
    for i in range(0,n_blenders):
        for j in range(0,n_output):
            Fixed_cost[i,j,'jp'] = 20+np.random.normal(mu, 4)
    for i in range(0,n_blenders):
        for j in range(0,n_blenders):
            Fixed_cost[i,j,'jj'] = 20+np.random.normal(mu, 4)

    Variable_cost = {}
    for i in range(0,n_streams):
        for j in range(0,n_blenders):
            Variable_cost[i,j,'sj'] = 5+np.random.normal(mu,1)
    for i in range(0,n_blenders):
        for j in range(0,n_output):
            Variable_cost[i,j,'jp'] = 5+np.random.normal(mu, 1)
    for i in range(0,n_blenders):
        for j in range(0,n_blenders):
            Variable_cost[i,j,'jj'] = 5+np.random.normal(mu, 1)
    data_config = {'n_streams':n_streams,'n_output':n_output,'n_properties':n_properties,'n_blenders':n_blenders,'n_periods':n_periods,
                               'supply':supply,'property':property,'lb_prop':lb_prop,'ub_prop':ub_prop,'inv_capacity':inv_capacity,
                               'Fixed_cost' : Fixed_cost,'Variable_cost':Variable_cost,'demand_price':demand_price}
    for key, value in data_config.items():
        globals()[key] = value
    #row_data = ["Checks"]
    #sheet.append(row_data)
    
    #row_data = ["Bayesopt"]
    #sheet.append(row_data)
    # Bounded region of parameter space

    model = ConcreteModel()
    model.C = Var(properties,blenders, points, domain=NonNegativeReals)
    model.Fsjt = Var(streams,blenders, periods, domain=NonNegativeReals)
    model.Fjpt = Var(blenders,output, periods, domain=NonNegativeReals)
    model.Ist = Var(streams, points, domain=NonNegativeReals)
    model.Ijt = Var(blenders, points, domain=NonNegativeReals)
    model.CI = Var(properties,blenders, points, domain=NonNegativeReals)
    model.CF = Var(properties,blenders,output, periods, domain=NonNegativeReals)
    model.Ipt = Var(output, points, domain=NonNegativeReals)
    model.pipt = Var(output, periods, domain=NonNegativeReals)
    def init_Is_rule(model, s):
        return model.Ist[s,0] == 0
    model.init_Is = Constraint(streams, rule=init_Is_rule)

    def init_Ip_rule(model, p):
        return model.Ipt[p,0] == 0
    model.init_Ip = Constraint(output, rule=init_Ip_rule)

    def init_Ij_rule(model, j):
        return model.Ijt[j,0] == 0
    model.init_Ij = Constraint(blenders, rule=init_Ij_rule)

    #def exec_fjjt(model, j,t):
        #return model.Fjjt[j,j,t] == 0
    #model.exec_same = Constraint(blenders, periods, rule=exec_fjjt)

    def balance_Is(model, s,t):
        return model.Ist[s,t] == model.Ist[s,t-1]-sum(model.Fsjt[s,j,t] for j in blenders)+supply[s,t]
    model.mb_Is = Constraint(streams, periods_mb, rule=balance_Is)

    def balance_Ij(model, j,t):
        return model.Ijt[j,t] == model.Ijt[j,t-1]+sum(model.Fsjt[s,j,t] for s in streams)-sum(model.Fjpt[j,p,t] for p in output)
    model.mb_Ij = Constraint(blenders, periods_mb, rule=balance_Ij)

    def balance_Ip(model, p,t):
        return model.Ipt[p,t] == model.Ipt[p,t-1]+sum(model.Fjpt[j,p,t] for j in blenders)-model.pipt[p,t]
    model.mb_Ip = Constraint(output, periods_mb, rule=balance_Ip)

    model.Xsjt = Var(streams,blenders, periods, domain=Binary)
    #model.Xjjt = Var(blenders,blenders, periods, domain=Binary)
    model.Xjpt = Var(blenders,output, periods, domain=Binary)
    #model.Xsjt.domain = Reals
    #model.Xjpt.domain = Reals

    
    M = 30
    def con_4_rule(model, s,j,t):
        return model.Fsjt[s,j,t]<=M*model.Xsjt[s,j,t]*param_val[3]
    model.con_4 = Constraint(streams,blenders,periods,rule=con_4_rule)

    def con_41_rule(model, j,p,t):
        return model.Fjpt[j,p,t]<=M*model.Xjpt[j,p,t]*param_val[4]
    model.con_41 = Constraint(blenders,output,periods,rule=con_41_rule)

    def con_51_rule(model, j,s,p,t):
        #return model.Ijt[j,t-1]>=sum(model.Fjpt[j,p,t] for p in output)
        return model.Xsjt[s,j,t]<=1-model.Xjpt[j,p,t]
    model.con_51 = Constraint(blenders,streams,output,periods,rule=con_51_rule)

    def con_53_rule(model, j,t):
        return model.Ijt[j,t-1]>=sum(model.Fjpt[j,p,t] for p in output)
        #return model.Xsjt[s,j,t]<=1-model.Xjpt[j,p,t]
    model.con_53 = Constraint(blenders,periods_mb,rule=con_53_rule)
    


    def balance_Cj(model, q,j,t):
    # return model.C[q,j,t]*model.Ijt[j,t] == model.Ijt[j,t-1]*model.C[q,j,t-1]+sum(model.Fsjt[s,j,t]*property[s,q] for s in streams)+sum(model.Fjjt[k,j,t]*model.C[q,k,t] for k in blenders)-sum(model.Fjjt[j,k,t]*model.C[q,j,t-1] for k in blenders)-sum(model.Fjpt[j,p,t]*model.C[q,j,t-1] for p in output)
        return model.CI[q,j,t] == model.CI[q,j,t-1]+sum(model.Fsjt[s,j,t]*property[s,q] for s in streams)-sum(model.CF[q,j,p,t] for p in output)
    model.mb_Cj = Constraint(properties,blenders, periods_mb, rule=balance_Cj)
    print(param_val)
    C_l = x4
    C_u = {}
    for q in properties:
        a = 0
        for p in output:
            a = max(a,ub_prop[p,q])
        C_u[q] = min(1,a)*param_val[0]
    
    I_l = 0
    I_l = 0
    I_u = max(inv_capacity['blender'])*param_val[1]
    F_l = 0
    F_u = M*param_val[2]
    def mckormik_1(m,q,j,t):
        return m.CI[q,j,t]>=C_l*m.Ijt[j,t]+m.C[q,j,t]*I_l-C_l*I_l
    model.mck1 = Constraint(properties,blenders,points,rule=mckormik_1)

    def mckormik_2(m,q,j,t):
        return m.CI[q,j,t]>=C_u[q]*m.Ijt[j,t]+m.C[q,j,t]*I_u-C_u[q]*I_u
    model.mck2 = Constraint(properties,blenders,points,rule=mckormik_2)

    def mckormik_3(m,q,j,t):
        return m.CI[q,j,t]<=C_l*m.Ijt[j,t]+m.C[q,j,t]*I_u-C_l*I_u
    model.mck3 = Constraint(properties,blenders,points,rule=mckormik_3)

    def mckormik_4(m,q,j,t):
        return m.CI[q,j,t]<=C_u[q]*m.Ijt[j,t]+m.C[q,j,t]*I_l-C_u[q]*I_l
    model.mck4 = Constraint(properties,blenders,points,rule=mckormik_4)

   
    def mckormik1_1(m,q,j,p,t):
        return m.CF[q,j,p,t]>=C_l*m.Fjpt[j,p,t]+m.C[q,j,t-1]*F_l-C_l*F_l
    model.mck11 = Constraint(properties,blenders,output,periods_mb,rule=mckormik1_1)

    def mckormik1_2(m,q,j,p,t):
        return m.CF[q,j,p,t]>=C_u[q]*m.Fjpt[j,p,t]+m.C[q,j,t-1]*F_u-C_u[q]*F_u
    model.mck12 = Constraint(properties,blenders,output,periods_mb,rule=mckormik1_2)

    def mckormik1_3(m,q,j,p,t):
        return m.CF[q,j,p,t]<=C_l*m.Fjpt[j,p,t]+m.C[q,j,t-1]*F_u-C_l*F_u
    model.mck13 = Constraint(properties,blenders,output,periods_mb,rule=mckormik1_3)

    def mckormik1_4(m,q,j,p,t):
        return m.CF[q,j,p,t]<=C_u[q]*m.Fjpt[j,p,t]+m.C[q,j,t-1]*F_l-C_u[q]*F_l
    model.mck14 = Constraint(properties,blenders,output,periods_mb,rule=mckormik1_4)

    def con_7_1rule(model, q,p,j,t):
        return lb_prop[p,q]*1-M*(1-model.Xjpt[j,p,t])*param_val[5] <= model.C[q,j,t-1]
    model.con_7_1 = Constraint(properties,output,blenders, periods_mb, rule=con_7_1rule)
    def con_7_2rule(model, q,p,j,t):
        return model.C[q,j,t-1]<=ub_prop[p,q]*1+M*(1-model.Xjpt[j,p,t])*param_val[6]
    model.con_7_2 = Constraint(properties,output,blenders, periods_mb, rule=con_7_2rule)

    def objective_rule(model):
        #return sum(sum(model.Xsjt[s,j,t]*Fixed_cost[s,j,'sj']+model.Fsjt[s,j,t]*Variable_cost[s,j,'sj'] for s in streams for j in blenders) for t in periods)+sum(sum(model.Xjpt[j,p,t]*Fixed_cost[j,p,'jp']+model.Fjpt[j,p,t]*Variable_cost[j,p,'jp'] for p in output for j in blenders) for t in periods) + sum(sum(model.Xjjt[j,k,t]*Fixed_cost[j,k,'jj']+model.Fjjt[j,k,t]*Variable_cost[j,k,'jj'] for k in blenders for j in blenders) for t in periods)
        return sum(sum(model.Xsjt[s,j,t]*Fixed_cost[s,j,'sj']*param_val[7]+model.Fsjt[s,j,t]*Variable_cost[s,j,'sj'] for s in streams for j in blenders) for t in periods)+sum(sum(model.Xjpt[j,p,t]*Fixed_cost[j,p,'jp']*param_val[8]+model.Fjpt[j,p,t]*Variable_cost[j,p,'jp'] for p in output for j in blenders) for t in periods)-sum(sum(model.pipt[p,t]*demand_price[p] for p in output) for t in periods)
    #model.obj = Objective(rule=objective_rule, sense=minimize)
    model.obj = Objective(rule=objective_rule, sense=minimize)

    solver = SolverFactory('gurobi',options = {'threads':8})  # 'couenne' is a solver for MINLP
    #solver.options['IIS'] = 1 
    results = solver.solve(model, tee=True)
    #model.computeIIS()
    
    
    if results.solver.termination_condition == TerminationCondition.infeasibleOrUnbounded:
        return -10**20
    actions = {}
    for t in periods:
        set = {}
        set['F_s_j'] = np.array([[model.Fsjt[s, j,t].value for j in blenders] for s in streams])
        #set['F_j_j'] = np.array([[model.Fjjt[j, k,t].value for k in blenders] for j in blenders])
        set['F_j_p'] = np.array([[model.Fjpt[j, p,t].value for p in output] for j in blenders])
        set['pi_p_t'] = np.array([model.pipt[p,t].value for p in output])
        actions[t-1] = set
    
    env_config = {'n_streams':n_streams,'n_output':n_output,'n_properties':n_properties,'n_blenders':n_blenders,'n_periods':n_periods,
                               'supply':supply,'demand_price':demand_price,'property':property,'lb_prop':lb_prop,'ub_prop':ub_prop,'inv_capacity':inv_capacity,
                               'Fixed_cost' : Fixed_cost,'Variable_cost':Variable_cost}
    #print(env_config)
    #print(env_config)
    env = Blending_env(config=env_config)
    
    state = env.reset()
    done = False
    rew_tot = 0
    while not done:
        time = env.current_step
        action = actions[time] # Take a random action based on the dynamic bounds
        #print(action)
        action_1 = np.concatenate([action['F_s_j'].flatten(),  action['F_j_p'].flatten(),action['pi_p_t'].flatten()])
        state, reward, done, info = env.step(action_1)
        rew_tot = rew_tot+reward
        env.render()
        print(reward)

    env.close()
    print(rew_tot)
    return rew_tot



def MBBF_val(x1,x2,x3,x4):
    val = 0
    n = 43
    for i in range(0,n):
        obj = instance_1(x1,x2,x3,x4)
        val = val+obj
    row_data = [x1,x2,x3,x4,val/n]
    sheet.append(row_data)
    
    workbook.save(file_path)
    return val
#MBBF(0.25156967213003645,1.07292020141285,1.2734394336887616)'''

pbounds = {'x1': (0, 1), 'x2': (0,1),'x3': (0, 1),'x4':(0,1)}

bo = BayesianOptimization(
        f=MBBF_val,
        pbounds=pbounds,
        random_state=1,
)

    
bo.maximize(
    init_points=2,
    n_iter=1000,
)
    #row_data = ["Checkbase"]
    #sheet.append(row_data)
    #MBBF(1.0,1.0,1.0,1.0,1.0,1.0,1.0)'''
#print(bo.max)
#MBBF_val(0.558857075,1,1,0.154025966)